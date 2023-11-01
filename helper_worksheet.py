import time
import openpyxl
from openpyxl.styles import Font

def prepare_worksheet_erros():
    workbook = openpyxl.Workbook()
    worksheet = workbook.active
    worksheet['A1'] = 'Page URL'
    worksheet['B1'] = 'Error Message'

    return (workbook, worksheet)

def prepare_worksheet_pagedata(url, page_title, meta_description, meta_og_title, meta_og_description):
    workbook = openpyxl.Workbook()
    worksheet = workbook.active

    worksheet.row_dimensions[7].height = 30
    worksheet.column_dimensions["A"].width = 22
    worksheet.column_dimensions["B"].width = 29
    worksheet.column_dimensions["C"].width = 29
    worksheet.column_dimensions["D"].width = 24
    worksheet.column_dimensions["E"].width = 24
    worksheet.column_dimensions["F"].width = 26

    worksheet.cell(row=1, column=1).value = 'URL:'
    worksheet.cell(row=2, column=1).value = 'Timestamp:'
    worksheet.cell(row=3, column=1).value = 'Title:'
    worksheet.cell(row=4, column=1).value = 'Meta Description:'
    worksheet.cell(row=5, column=1).value = 'Meta OG Title:'
    worksheet.cell(row=6, column=1).value = 'Meta OG Description:'

    worksheet.cell(row=1, column=2).value = url
    worksheet.cell(row=2, column=2).value = time.strftime("%Y-%m-%d %H:%M:%S")
    worksheet.cell(row=3, column=2).value = page_title
    worksheet.cell(row=4, column=2).value = meta_description
    worksheet.cell(row=5, column=2).value = meta_og_title
    worksheet.cell(row=6, column=2).value = meta_og_description

    worksheet.cell(row=7, column=1).value = 'Tag'
    worksheet.cell(row=7, column=2).value = 'Content'
    worksheet.cell(row=7, column=3).value = 'Filename'
    worksheet.cell(row=7, column=4).value = 'Alt Text'
    worksheet.cell(row=7, column=5).value = 'Title Text'
    worksheet.cell(row=7, column=6).value = 'Media Text'

    for cell in worksheet["A1:A5"][0] + worksheet["A6:F6"][0]:
        cell.font = Font(bold=True)

    return (workbook, worksheet)

def write_worksheet_pagedata(worksheet, tag_name, content, file_name, alt_text, title_text, media_text):
    next_row = len(worksheet['A']) + 1

    worksheet.cell(row=next_row, column=1, value=tag_name)
    worksheet.cell(row=next_row, column=1).font = Font(bold=True)
    worksheet.cell(row=next_row, column=2, value=content)
    worksheet.cell(row=next_row, column=3, value=file_name)
    worksheet.cell(row=next_row, column=4, value=alt_text)
    worksheet.cell(row=next_row, column=5, value=title_text)
    worksheet.cell(row=next_row, column=6, value=media_text)
